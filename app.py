from email.policy import default
import os
import csv
from datetime import datetime
import io
from flask import Flask, render_template, request, redirect, url_for, flash, make_response
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text, func
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO

# ============================================================================
# CONFIGURACIÓN DE LA APLICACIÓN
# ============================================================================
app = Flask(__name__)

# Clave secreta para sesiones y mensajes flash (CRÍTICO: usar variable de entorno en producción)
app.secret_key = os.environ.get('SECRET_KEY', 'mi-clave-super-larga-y-compleja-2024')

# Configuración de base de datos
# En Render usa PostgreSQL (DATABASE_URL), localmente usa SQLite
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
    "DATABASE_URL",
    "sqlite:///novedades.db"
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Inicializar SQLAlchemy
db = SQLAlchemy(app)


# ============================================================================
# MODELO DE BASE DE DATOS
# ============================================================================
class Novedad(db.Model):
    """Modelo que representa una novedad laboral (Alta/Baja/Reemplazo/Otros)"""
    __tablename__ = 'novedades'

    # ID y timestamp
    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.String(50), nullable=False)

    # Datos básicos del empleado
    legajo = db.Column(db.String(50), default="")
    nombre = db.Column(db.String(255), nullable=False)
    tipo_empleado = db.Column(db.String(50), default="")  # Docente/No Docente
    tipo_novedad = db.Column(db.String(50), nullable=False)  # Alta/Baja/Reemplazo/Otros

    # Datos personales
    fecha_nacimiento = db.Column(db.String(50), default="")
    dni = db.Column(db.String(50), default="")
    cuil = db.Column(db.String(50), default="")
    cbu = db.Column(db.String(50), default="")
    banco = db.Column(db.String(255), default="")
    domicilio = db.Column(db.String(255), default="")
    email = db.Column(db.String(255), default="")
    obra_social = db.Column(db.String(255), default="")

    # Datos del alta
    nivel = db.Column(db.String(50), default="")  # Inicial/Primario/Secundario/Terciario
    fecha_alta = db.Column(db.String(50), default="")
    cargo = db.Column(db.String(255), default="")
    caracter_del_cargo = db.Column(db.String(50), default="")  # Titular/Suplente

    # Decretos
    trabaja_otra_institucion = db.Column(db.String(50), default="")
    tipo_institucion = db.Column(db.String(50), default="")  # Pública/Privada/Ambas
    horas_catedras = db.Column(db.Float, default=0.0)

    # Datos laborales
    subvencionado = db.Column(db.String(50), default="")
    asignaciones_familiares = db.Column(db.String(50), default="")
    cantidad_hijos = db.Column(db.Integer, default=0)

    # Datos de reemplazo
    reemplazo_persona_ya_trabaja = db.Column(db.String(50), default="")
    reemplazo_cargo_que_cubre = db.Column(db.String(255), default="")
    fecha_inicio_reemplazo = db.Column(db.String(50), default="")
    fecha_fin_reemplazo = db.Column(db.String(50), default="")

    # Datos de baja
    fecha_baja = db.Column(db.String(50), default="")
    motivo_baja = db.Column(db.String(255), default="")

    # Otros tipos de novedad
    tipo_otro = db.Column(db.String(255), default="")  # Anticipo/Inasistencia/Licencia

    # Información adicional
    cargos_actuales = db.Column(db.String(255), default="")
    tipo_movimiento = db.Column(db.String(255), default="")
    subvencion = db.Column(db.String(255), default="")
    codigo = db.Column(db.String(255), default="")
    observaciones = db.Column(db.Text, default="")


# ============================================================================
# FUNCIONES DE BASE DE DATOS
# ============================================================================
def add_novedad(data: dict):
    """
    Guarda una nueva novedad en la base de datos.
    
    Args:
        data: Diccionario con los campos de la novedad
    """
    nueva_novedad = Novedad(**data)
    db.session.add(nueva_novedad)
    db.session.commit()


def get_all_novedades():
    """
    Obtiene todas las novedades del período actual (del 6 del mes al 5 del siguiente).
    
    Returns:
        Lista de objetos Novedad ordenados por fecha descendente
    """
    today = date.today()

    # Calcular el rango de fechas (del 6 al 5)
    if today.day >= 6:
        # Desde el 6 de este mes hasta el 5 del mes siguiente
        start_date = today.replace(day=6)
        end_date = (today.replace(day=1) + relativedelta(months=1)).replace(day=5)
    else:
        # Desde el 6 del mes anterior hasta el 5 de este mes
        start_date = (today.replace(day=1) - relativedelta(months=1)).replace(day=6)
        end_date = today.replace(day=5)

    # Consultar base de datos con filtro de fechas
    novedades = (
        Novedad.query
        .filter(func.date(Novedad.timestamp) >= start_date)
        .filter(func.date(Novedad.timestamp) <= end_date)
        .order_by(Novedad.timestamp.desc())
        .all()
    )

    return novedades

def row_to_dict(obj):
    """Convierte un objeto SQLAlchemy en un diccionario limpio"""
    return {c.name: getattr(obj, c.name) for c in obj.__table__.columns}



# ============================================================================
# RUTAS DE LA APLICACIÓN
# ============================================================================
@app.route("/", methods=["GET"])
def index():
    """Página principal con el formulario de carga o edición"""
    #return render_template("index.html")
    edit_id = request.args.get("edit_id")

    if edit_id:
        #Buscar registro en la BD
        novedad = Novedad.query.get(int(edit_id))

        if not novedad:
            flash("No se encontró la novedad para editar.", "danger")
            return render_template("index.html", edit_mode=False)
        
        flash("Editando novedad existente", "info")
        return render_template("index.html", edit_mode=True, data=row_to_dict(novedad))
    #Modo normal (alta)
    return render_template("index.html", edit_mode=False, data={})
    



@app.route("/enviar", methods=["POST"])
def enviar():
    """Procesa y valida el formulario de novedades"""
    
    # -------------------------------------------------------------------------
    # 1. OBTENER DATOS DEL FORMULARIO
    # -------------------------------------------------------------------------
    # Datos básicos
    legajo = (request.form.get("legajo") or "").strip()
    nombre = (request.form.get("nombre") or "").strip()
    tipo_empleado = (request.form.get("tipo_empleado") or "").strip()
    tipo_novedad = (request.form.get("tipo_novedad") or "").strip()
    
    # Datos personales
    fecha_nacimiento = (request.form.get("fecha_nacimiento") or "").strip()
    dni = (request.form.get("dni") or "").strip()
    cuil = (request.form.get("cuil") or "").strip()
    cbu = (request.form.get("cbu") or "").strip()
    banco = (request.form.get("banco") or "").strip()
    domicilio = (request.form.get("domicilio") or "").strip()
    email = (request.form.get("email") or "").strip()
    obra_social = (request.form.get("obra_social") or "").strip()
    
    # Datos del alta
    nivel = (request.form.get("nivel") or "").strip()
    fecha_alta = (request.form.get("fecha_alta") or "").strip()
    cargo = (request.form.get("cargo") or "").strip()
    caracter_del_cargo = (request.form.get("caracter_del_cargo") or "").strip()
    
    # Decretos
    trabaja_otra_institucion = (request.form.get("trabaja_otra_institucion") or "").strip()
    tipo_institucion = (request.form.get("tipo_institucion") or "").strip()
    horas_catedras_str = (request.form.get("horas_catedras") or "").strip()
    
    # Datos laborales
    subvencionado = (request.form.get("subvencionado") or "").strip()
    asignaciones_familiares = (request.form.get("asignaciones_familiares") or "").strip()
    cantidad_hijos_str = (request.form.get("cantidad_hijos") or "").strip()
    
    # Datos de reemplazo
    reemplazo_persona_ya_trabaja = (request.form.get("reemplazo_persona_ya_trabaja") or "").strip()
    reemplazo_cargo_que_cubre = (request.form.get("reemplazo_cargo_que_cubre") or "").strip()
    fecha_inicio_reemplazo = (request.form.get("fecha_inicio_reemplazo") or "").strip()
    fecha_fin_reemplazo = (request.form.get("fecha_fin_reemplazo") or "").strip()
    
    # Datos de baja
    fecha_baja = (request.form.get("fecha_baja") or "").strip()
    motivo_baja = (request.form.get("motivo_baja") or "").strip()
    
    # Otros
    tipo_otro = (request.form.get("tipo_otro") or "").strip()
    
    # Información adicional
    cargos_actuales = (request.form.get("cargos_actuales") or "").strip()
    tipo_movimiento = (request.form.get("tipo_movimiento") or "").strip()
    subvencion = (request.form.get("subvencion") or "").strip()
    codigo = (request.form.get("codigo") or "").strip()
    observaciones = (request.form.get("observaciones") or "").strip()



#@app.route("/actualizar", methods=["POST"])
#def actualizar():
#    """Actualiza una novedad existente"""
    
#    id_str = request.form.get("id")
#    if not id_str:
#        flash("Falta el ID de la novedad a editar.", "danger")
#        return redirect(url_for("ver"))

#    novedad = Novedad.query.get(int(id_str))
#    if not novedad:
#        flash("No se encontró la novedad a editar.", "danger")
#        return redirect(url_for("ver"))
    
    # -------------------------------------------------------------------------
    # ACTUALIZAR TODOS LOS CAMPOS (igual que en /enviar)
    # -------------------------------------------------------------------------


    # Datos personales
#    fecha_nacimiento = (request.form.get("fecha_nacimiento") or "").strip()
#    novedad.fecha_nacimiento = fecha_nacimiento

# --------- ACTUALIZAR CAMPOS (SIN tocar id ni timestamp) ---------
    #updates = {
        
    
     #   "fecha_nacimiento": fecha_nacimiento,
    #}

    #for k, v in updates.items():
    #setattr(novedad, k, v)


    # -------------------------------------------------------------------------
    # 2. VALIDACIONES
    # -------------------------------------------------------------------------
    errores = []
    
    # Validaciones básicas
    if not nombre:
        errores.append("El campo 'Nombre y apellido' es obligatorio.")
    if not tipo_empleado:
        errores.append("Seleccioná 'Docente' o 'No Docente'.")
    if not tipo_novedad:
        errores.append("Seleccioná el 'Tipo de novedad'.")
    
    ## Validaciones por tipo de novedad
    if tipo_novedad == "Alta":
         if not fecha_alta or not nivel or not cargo:
            errores.append("Para 'Alta', los campos 'Fecha de alta', 'Nivel' y 'Cargo' son obligatorios.")
    
    elif tipo_novedad == "Baja":
        if not fecha_baja or not motivo_baja:
            errores.append("Para 'Baja', los campos 'Fecha de baja' y 'Motivo' son obligatorios.")
    
    elif tipo_novedad == "Reemplazo":
        if reemplazo_persona_ya_trabaja not in ("Si", "No"):
            errores.append("En 'Reemplazo', indicá si la persona ya trabaja en el colegio (Sí/No).")
        elif reemplazo_persona_ya_trabaja == "Si":
            if not reemplazo_cargo_que_cubre or not fecha_inicio_reemplazo or not fecha_fin_reemplazo:
                errores.append("En 'Reemplazo (ya trabaja)', indicá cargo y fechas de inicio y fin.")
        elif not fecha_alta or not nivel or not cargo:
            errores.append("En 'Reemplazo (nuevo)', los campos 'Fecha de alta', 'Nivel' y 'Cargo' son obligatorios.")
    
    elif tipo_novedad == "Otros":
        if not tipo_otro:
            errores.append("En 'Otros', seleccioná el subtipo (Anticipo/Inasistencia/Lic. sin goce).")

    # Validación de números
    try:
        horas_catedras = float(horas_catedras_str.replace(",", ".")) if horas_catedras_str else 0.0
    except ValueError:
        errores.append("Horas Cátedras debe ser un número válido.")
        horas_catedras = 0.0

    try:
        cantidad_hijos = int(cantidad_hijos_str) if cantidad_hijos_str else 0
    except ValueError:
        errores.append("Cantidad de hijos debe ser un número entero.")
        cantidad_hijos = 0

    # Si hay errores, mostrarlos y volver al formulario
    if errores:
        for e in errores:
            flash(e, "danger")
        return render_template("index.html", edit_mode=False, data=request.form.to_dict())

    # -------------------------------------------------------------------------
    # 3. CREAR REGISTRO Y GUARDAR
    # -------------------------------------------------------------------------
    registro = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "legajo": legajo,
        "nombre": nombre,
        "tipo_empleado": tipo_empleado,
        "tipo_novedad": tipo_novedad,
        "fecha_nacimiento": fecha_nacimiento,
        "dni": dni,
        "cuil": cuil,
        "cbu": cbu,
        "banco": banco,
        "domicilio": domicilio,
        "email": email,
        "obra_social": obra_social,
        "nivel": nivel,
        "fecha_alta": fecha_alta,
        "cargo": cargo,
        "caracter_del_cargo": caracter_del_cargo,
        "trabaja_otra_institucion": trabaja_otra_institucion,
        "tipo_institucion": tipo_institucion,
        "horas_catedras": horas_catedras,
        "subvencionado": subvencionado,
        "asignaciones_familiares": asignaciones_familiares,
        "cantidad_hijos": cantidad_hijos,
        "reemplazo_persona_ya_trabaja": reemplazo_persona_ya_trabaja,
        "reemplazo_cargo_que_cubre": reemplazo_cargo_que_cubre,
        "fecha_inicio_reemplazo": fecha_inicio_reemplazo,
        "fecha_fin_reemplazo": fecha_fin_reemplazo,
        "fecha_baja": fecha_baja,
        "motivo_baja": motivo_baja,
        "tipo_otro": tipo_otro,
        "cargos_actuales": cargos_actuales,
        "tipo_movimiento": tipo_movimiento,
        "subvencion": subvencion,
        "codigo": codigo,
        "observaciones": observaciones,
    }

    add_novedad(registro)
    flash("¡Novedad registrada correctamente!", "success")
    return redirect(url_for("ver"))


@app.route("/actualizar", methods=["POST"])
def actualizar():
    """Actualiza una novedad existente (por ahora solo fecha_nacimiento)."""
    id_str = request.form.get("id")
    if not id_str:
        flash("Falta el ID de la novedad a editar.", "danger")
        return redirect(url_for("ver"))

    novedad = Novedad.query.get(int(id_str))
    if not novedad:
        flash("No se encontró la novedad a editar.", "danger")
        return redirect(url_for("ver"))

    # ---- SOLO ACTUALIZAMOS FECHA DE NACIMIENTO (ejemplo) ----
    novedad.fecha_nacimiento = (request.form.get("fecha_nacimiento") or "").strip()
    novedad.nombre = (request.form.get("nombre") or "").strip()

    db.session.commit()
    flash("¡Novedad actualizada!", "success")
    return redirect(url_for("ver"))






@app.route("/ver", methods=["GET"])
def ver():
    """Muestra la tabla con todas las novedades del período actual"""
    novedades_db = get_all_novedades()
    
    # Convertir objetos a diccionarios para el template
    rows = [n.__dict__ for n in novedades_db]
    
    return render_template("ver.html", rows=rows, total=len(rows))


# REEMPLAZAR la función descargar() completa por esta:
@app.route("/descargar", methods=["GET"])
def descargar():
    """Descarga las novedades en formato XLSX profesional"""
    
    # Obtener todas las novedades
    novedades_db = get_all_novedades()
    
    # Validar que hay datos
    if not novedades_db:
        flash("No hay datos para descargar.", "warning")
        return redirect(url_for("ver"))
    
    # -------------------------------------------------------------------------
    # PREPARAR DATOS
    # -------------------------------------------------------------------------
    def format_date(date_str):
        """Convierte fechas de YYYY-MM-DD a DD/MM/YYYY"""
        if not date_str:
            return ""
        try:
            if len(date_str) == 10 and date_str[4] == '-':
                parts = date_str.split('-')
                return f"{parts[2]}/{parts[1]}/{parts[0]}"
            return date_str
        except:
            return date_str
    
    def format_number(num_value):
        """Convierte números con punto decimal a coma"""
        if num_value is None or num_value == 0 or num_value == 0.0:
            return ""
        return str(num_value).replace('.', ',')
    
    def format_timestamp(timestamp_str):
        """Extrae solo la fecha del timestamp"""
        if not timestamp_str:
            return ""
        try:
            date_part = timestamp_str.split('T')[0] if 'T' in timestamp_str else timestamp_str.split(' ')[0]
            return format_date(date_part)
        except:
            return timestamp_str
    
    # -------------------------------------------------------------------------
    # CREAR ESTRUCTURA DE DATOS
    # -------------------------------------------------------------------------
    rows_data = []
    for novedad in novedades_db:
        row = {
            'Fecha Carga': format_timestamp(novedad.timestamp),
            'Legajo': novedad.legajo or "",
            'Nombre y Apellido': novedad.nombre or "",
            'Tipo Empleado': novedad.tipo_empleado or "",
            'Tipo Novedad': novedad.tipo_novedad or "",
            'Fecha Nacimiento': format_date(novedad.fecha_nacimiento),
            'DNI': novedad.dni or "",
            'CUIL': novedad.cuil or "",
            'CBU': novedad.cbu or "",
            'Banco': novedad.banco or "",
            'Domicilio': novedad.domicilio or "",
            'Email': novedad.email or "",
            'Obra Social': novedad.obra_social or "",
            'Nivel': novedad.nivel or "",
            'Fecha Alta': format_date(novedad.fecha_alta),
            'Cargo': novedad.cargo or "",
            'Carácter Cargo': novedad.caracter_del_cargo or "",
            'Trabaja Otra Inst': novedad.trabaja_otra_institucion or "",
            'Tipo Institución': novedad.tipo_institucion or "",
            'Horas Cátedras': format_number(novedad.horas_catedras),
            'Subvencionado': novedad.subvencionado or "",
            'Asig. Familiares': novedad.asignaciones_familiares or "",
            'Cantidad Hijos': str(novedad.cantidad_hijos) if novedad.cantidad_hijos else "",
            'Ya Trabaja': novedad.reemplazo_persona_ya_trabaja or "",
            'Cargo que Cubre': novedad.reemplazo_cargo_que_cubre or "",
            'Inicio Reemplazo': format_date(novedad.fecha_inicio_reemplazo),
            'Fin Reemplazo': format_date(novedad.fecha_fin_reemplazo),
            'Fecha Baja': format_date(novedad.fecha_baja),
            'Motivo Baja': novedad.motivo_baja or "",
            'Tipo Otro': novedad.tipo_otro or "",
            'Cargos Actuales': novedad.cargos_actuales or "",
            'Tipo Movimiento': novedad.tipo_movimiento or "",
            'Subvención': novedad.subvencion or "",
            'Código': novedad.codigo or "",
            'Observaciones': novedad.observaciones or ""
        }
        rows_data.append(row)
    
    # -------------------------------------------------------------------------
    # CREAR ARCHIVO EXCEL
    # -------------------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Novedades"
    
    # Headers (nombres de columnas)
    headers = list(rows_data[0].keys())
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Escribir encabezados con estilo
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Escribir datos
    for row_num, row_data in enumerate(rows_data, 2):
        for col_num, header in enumerate(headers, 1):
            value = row_data[header]
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical='center')
    
    # Ajustar ancho de columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Congelar primera fila (encabezados)
    ws.freeze_panes = 'A2'
    
    # -------------------------------------------------------------------------
    # GUARDAR Y ENVIAR ARCHIVO
    # -------------------------------------------------------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nombre del archivo con fecha actual
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    filename = f"novedades_{fecha_actual}.xlsx"
    
    # Crear respuesta
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    return response


# ============================================================================
# INICIALIZACIÓN
# ============================================================================
@app.before_request
def create_tables():
    """Crea las tablas en la base de datos si no existen"""
    try:
        with app.app_context():
            db.create_all()
            print("✓ Tablas de la base de datos verificadas")
    except Exception as e:
        print(f"✗ Error al crear tablas: {e}")


if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    app.run(debug=True)